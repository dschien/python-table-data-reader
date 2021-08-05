from pathlib import Path
from typing import List, Dict
from collections import namedtuple
from openpyxl import load_workbook
from table_data_reader.table_handlers import OpenpyxlTableHandler

import logging

logger = logging.getLogger(__name__)

Cell_value = namedtuple('Cell_value', ['cell_ref', 'value'])


class TableWriter(OpenpyxlTableHandler):

    def __init__(self, workbook_input_path: Path = None, worksheets: List = None, workbook_output_path: Path = None):
        """

        :type workbook_input_path: Path object
        """
        super().__init__()
        self.worksheets = worksheets
        self.workbook_path = workbook_input_path
        self.workbook_output_path = workbook_output_path if workbook_output_path else workbook_input_path
        self.workbook = load_workbook(workbook_input_path)

    def update_table(self, data: List[Dict[str, float]]) -> None:
        """
        Iterate over all cells, if the id is identical, then overwrite the ref value provided.
        Overwrites the table

        :param data:
        :return:
        """
        _data = {item['id']: item['value'] for item in data}

        def update_row_visitor(wb=None, row=None, header=None, **kwargs):
            cell_map = {}
            for key, cell in zip(header, row):
                cell_map[key] = Cell_value(cell_ref=cell, value=cell.value)
            var_id = cell_map['id'].value
            if var_id in _data.keys():
                # handles writing group variables
                # todo: this won't work with group variables that are on the main sheet
                if isinstance(_data[var_id], dict):
                    # if a sheet for the variable already exists
                    if cell_map['variable'].value in wb.sheetnames:
                        sheet = wb[cell_map['variable'].value]

                        scenario = cell_map['scenario'].value

                        # get the header row for the variable sheet
                        rows = list(sheet.iter_rows())
                        variable_sheet_headers = [cell.value for cell in rows[0]]
                        ref_column = variable_sheet_headers.index('ref value')

                        written_groups = set()

                        groups = list(_data[var_id].keys())

                        # go through existing rows to see if there is something to override
                        for row in rows[1:]:
                            row_scenario = row[variable_sheet_headers.index('scenario')].value
                            if row_scenario == scenario:
                                group = row[variable_sheet_headers.index('group')].value
                                if group in groups:
                                    # if the scenario and group match, override the ref value and mark as written
                                    row[ref_column].value = _data[var_id][group]
                                    written_groups.add(group)

                        # go through groups that havent been overwritten, and add them to the end of the sheet
                        row_index = len(rows) + 1
                        for group in groups:
                            if group not in written_groups:
                                for i, header in enumerate(variable_sheet_headers):
                                    if header == 'id':
                                        continue
                                    if header in cell_map.keys():
                                        sheet.cell(row=row_index, column=i + 1).value = cell_map[header].value

                                sheet.cell(row=row_index, column=1).value = group
                                sheet.cell(row=row_index, column=ref_column + 1).value = _data[var_id][group]

                                row_index += 1

                    else:
                        sheet = wb.create_sheet(cell_map['variable'].value)
                        logger.info(f'Creating group sheet for variable {cell_map["variable"]} (id {var_id})')
                        variable_sheet_headers = ['group', 'scenario', 'ref value', 'mean growth',
                                                  'initial_value_proportional_variation', 'variability growth', 'id']
                        ref_column = variable_sheet_headers.index('ref value') + 1

                        for i, header in enumerate(variable_sheet_headers):
                            sheet.cell(row=1, column=i+1).value = header

                        groups = list(_data[var_id].keys())
                        for i, group in enumerate(groups):
                            for j, header in enumerate(variable_sheet_headers):
                                if header == 'id':
                                    continue
                                if header in cell_map.keys():
                                    sheet.cell(row=i+2, column=j+1).value = cell_map[header].value

                            sheet.cell(row=i+2, column=1).value = group
                            sheet.cell(row=i+2, column=ref_column).value = _data[var_id][group]
                else:
                    # write the updated value into cell
                    logger.info(
                        f'Overwriting template value for variable {cell_map["variable"]} (id {var_id}) with new value {_data[var_id]} (was {cell_map["ref value"].value})')
                    cell_map['ref value'].cell_ref.value = _data[var_id]

        self.table_visitor(wb=self.workbook, sheet_names=self.worksheets, visitor_function=update_row_visitor)
        logger.info(f'Writing updated workbook to file {self.workbook_output_path}')
        self.workbook.save(self.workbook_output_path)
        self.workbook.close()
