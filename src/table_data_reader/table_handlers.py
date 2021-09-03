import csv
from abc import abstractmethod
from collections import defaultdict
import datetime
from numbers import Number
from typing import Dict, List
from functools import partial

import openpyxl
from openpyxl import Workbook
from typing import Callable

import json

import logging

logger = logging.getLogger(__name__)

from table_data_reader import param_name_maps, ParameterRepository, Parameter


class TableValidationError(ValueError):
    pass


class TableHandler(object):
    version: int

    def __init__(self, version=2):
        self.version = version

    @abstractmethod
    def load_definitions(self, sheet_name, filename=None, id_flag=False, **kwargs):  # pragma: no cover
        raise NotImplementedError()


class Xlsx2CsvHandler(TableHandler):
    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        from xlsx2csv import Xlsx2csv
        data = Xlsx2csv(filename, inmemory=True).convert(None, sheetid=0)

        definitions = []

        _sheet_names = [sheet_name] if sheet_name else [data.keys()]

        for _sheet_name in _sheet_names:
            sheet = data[_sheet_name]

            header = sheet.header
            if header[0] != 'variable':
                continue

            for row in sheet.rows:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell
                definitions.append(values)
        return definitions


class DictReaderStrip(csv.DictReader):
    @property
    def fieldnames(self):
        if self._fieldnames is None:
            # Initialize self._fieldnames
            # Note: DictReader is an old-style class, so can't use super()
            csv.DictReader.fieldnames.fget(self)
            if self._fieldnames is not None:
                self._fieldnames = [name.strip() for name in self._fieldnames]
        return self._fieldnames


class CSVHandler(TableHandler):
    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        reader = DictReaderStrip(open(filename), delimiter=',')

        definitions = []

        _definition_tracking = defaultdict(dict)

        for i, row in enumerate(reader):

            values = {k: v.strip() for k, v in row.items()}

            if not values['variable']:
                logger.debug(f'ignoring row {i}: {row}')
                continue
            for key in ['ref value', 'initial_value_proportional_variation', 'mean growth', 'variability growth']:
                try:
                    new_val = float(values[key])
                    values[key] = new_val
                except:
                    if values['type'] == 'interp':
                        continue
                    else:
                        raise Exception(
                            f'Could not convert value <{values[key]}> for key {key} to number in row {i} for variable {values["variable"]}')

            if 'ref date' in values and values['ref date']:
                if isinstance(values['ref date'], str):
                    values['ref date'] = datetime.datetime.strptime(values['ref date'], '%d/%m/%Y')
                    if values['ref date'].day != 1:
                        logger.warning(
                            f'ref date truncated to first of month for variable {values["variable"]}')
                        values['ref date'] = values['ref date'].replace(day=1)
                else:
                    raise Exception(
                        f"{values['ref date']} for variable {values['variable']} is not a date - "
                        f"check spreadsheet value is a valid day of a month")
            logger.debug(f'values for {values["variable"]}: {values}')
            definitions.append(values)
            scenario = values['scenario'] if values['scenario'] else "n/a"

            if scenario in _definition_tracking[values['variable']]:

                logger.error(
                    f"Duplicate entry for parameter "
                    f"with name <{values['variable']}> and <{scenario}> scenario in file")
                raise ValueError(
                    f"Duplicate entry for parameter "
                    f"with name <{values['variable']}> and <{scenario}> scenario in file")

            else:
                _definition_tracking[values['variable']][scenario] = 1
        return definitions


class PandasCSVHandler(TableHandler):

    def strip(self, text):
        try:
            return text.strip()
        except AttributeError:
            return text

    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        self.version = 2

        import pandas as pd
        df = pd.read_csv(filename, usecols=range(15), index_col=False, parse_dates=['ref date'],
                         dtype={'initial_value_proportional_variation': 'float64'},
                         dayfirst=True,
                         # date_parser=l0ambda x: pd.datetime.strptime(x, '%d-%m-%Y')
                         )
        df = df.dropna(subset=['variable', 'ref value'])
        df.fillna("", inplace=True)

        return df.to_dict(orient='records')


class OpenpyxlTableHandler(TableHandler):
    version: int

    def __init__(self, version=2):
        super().__init__(version=version)

    def group_builder(self, entry: Dict = None, group_variables=None, sheet_name=None, **kwargs):
        """
        Mutates the group_variables dictionary to store group-level variable values
        Dictionary is organised as dict[variable][scenario][group]
        :param entry:
        :param group_variables:
        :param sheet_name:
        :param kwargs:
        :return:
        """

        var = entry["variable"]
        group = entry.get("group", None)
        scenario = entry["scenario"] if entry.get("scenario", None) else "default"
        if group is not None:
            if var not in group_variables.keys():
                group_variables[var] = {}
            if scenario not in group_variables[var].keys():
                group_variables[var][scenario] = {}
            if group in group_variables[var][scenario].keys():
                logger.error(
                    f"Duplicate entry for parameter "
                    f"with name <{var}>,<{group}> scenario, and <{scenario}> group in sheet {sheet_name}")
                raise ValueError(
                    f"Duplicate entry for parameter "
                    f"with name <{var}>,<{group}> scenario, and <{scenario}> group in sheet {sheet_name}")
            group_variables[var][scenario][group] = entry

    def truncate_ref_date(self, values: Dict = None):
        """
        Truncates ref dates to the beginning of the month
        """

        if values.get('ref date') is not None:
            if isinstance(values['ref date'], datetime.datetime):
                if values['ref date'].day != 1:
                    logger.warning(f'ref date truncated to first of month for variable {values["variable"]}')
                    values['ref date'] = values['ref date'].replace(day=1)
            else:
                raise Exception(
                    f"{values['ref date']} for variable {values['variable']} is not a date - "
                    f"check spreadsheet value is a valid day of a month")
        return values

    def build_definitions(self, entry: Dict = None, definitions=None, sheet_name=None,
                          group_flag=False, group_variables=None, wb=None, **kwargs):
        """
        Assigns group-level dictionaries to parameter values in definitions with weird dictionary stuff
        :param entry:
        :param definitions:
        :param sheet_name:
        :param group_flag:
        :param group_variables:
        :param wb:
        :param kwargs:
        :return:
        """

        entry = self.truncate_ref_date(entry)

        logger.debug(f'values for {entry["variable"]}: {entry}')
        variable_name = entry['variable']
        scenario = entry['scenario'] if entry.get('scenario', None) else "default"

        if scenario in definitions[variable_name].keys():
            # if this is an inline group row the error doesn't need to be raised as it's normal
            if entry.get('group', None) is not None:
                return None
            logger.error(
                f"Duplicate entry for parameter "
                f"with name <{entry['variable']}> and <{scenario}> scenario in sheet {sheet_name}")
            raise ValueError(
                f"Duplicate entry for parameter "
                f"with name <{entry['variable']}> and <{scenario}> scenario in sheet {sheet_name}")
        else:
            # if the group flag is not on or there is no sheet by this parameter name just read from params
            if not group_flag or (variable_name not in wb.sheetnames and variable_name not in group_variables.keys()):
                definitions[variable_name][scenario] = entry
            else:
                keys = list(entry.keys())
                group_values = {}

                # set parameters that should be constant across each subvariable in the group to the same value
                group_constants = ["variable", "type", "param", "unit"]
                if 'scenario' in keys:
                    group_constants.append('scenario')
                if 'group' in keys:
                    group_constants.append('group')
                for group_constant in group_constants:
                    keys.remove(group_constant)
                    group_values[group_constant] = entry[group_constant]

                for key in keys:
                    group_values[key] = {}
                if variable_name in group_variables.keys():
                    # we have already parsed this group variable in inline_groupings
                    # so just set group_values here
                    scenarios = group_variables[variable_name].keys()
                    if scenario in scenarios:
                        groups = group_variables[variable_name][scenario].keys()
                        for group in groups:
                            for key in keys:
                                value = group_variables[variable_name][scenario][group][key]

                                if value is not None:
                                    group_values[key][group] = value
                                else:
                                    group_values[key][group] = entry[key]
                else:
                    # the variable is a group variable but has not been parsed inline as part of the main page
                    # so, find its sheet and read from it.
                    # todo: move this into groupings_handler?
                    rows = list(wb[variable_name].iter_rows())
                    header = [cell.value for cell in rows[0]]
                    for i, row in enumerate(rows[1:]):
                        temp_values = {}
                        for key, cell in zip(header, row):
                            temp_values[key] = cell.value  # reads values from the variable's sheet
                        temp_scenario = temp_values['scenario'] if temp_values['scenario'] else "default"
                        if temp_scenario == scenario:
                            for key in keys:
                                if key in header and temp_values[key] is not None:
                                    group_values[key][temp_values["group"]] = temp_values[key]
                                else:
                                    group_values[key][temp_values["group"]] = entry[key]

                ref_dates = list(group_values['ref date'].values())
                # Ensures that every element in ref_dates is the same
                # todo: see if we can remove this restriction
                assert ref_dates.count(ref_dates[0]) == len(ref_dates), \
                    f"Different groups have different ref dates for {entry['variable']}"
                group_values['ref date'] = ref_dates[0]

                definitions[variable_name][scenario] = group_values

    def table_visitor(self, wb: Workbook = None, sheet_names: List[str] = None, visitor_function: Callable = None,
                      definitions=None, **kwargs):
        """
        stub for id management

        todo: try and remove checks for specific kwargs to allow for more generic visitor functionality

        todo: make this work for other tables? might not be worth the effort

        :param definitions:
        :param wb:
        :type wb:
        :param sheet_names:
        :type sheet_names:
        :param visitor_function:
        :type visitor_function:
        :return:
        :rtype:
        """
        if not sheet_names:
            sheet_names = wb.sheetnames
        for _sheet_name in sheet_names:
            if _sheet_name == 'metadata':
                continue
            sheet = wb[_sheet_name]
            rows = list(sheet.iter_rows())
            header = [cell.value for cell in rows[0]]
            if header[0] != 'variable':
                continue
            for i, row in enumerate(rows[1:]):
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                if not values['variable']:
                    logger.debug(f'ignoring row {i}: {row}')
                    continue

                group_flag = kwargs.get('with_group') and (values['variable'] in kwargs['group_vars'])

                visitor_function(ws=sheet, entry=values, definitions=definitions,
                                 row_idx=i, sheet_name=_sheet_name, row=row,
                                 header=header, wb=wb, group_flag=group_flag, **kwargs)
        return definitions

    def get_version(self, wb):
        # todo: test versioning?
        version = 1
        try:
            sheet = wb['metadata']
            rows = list(sheet.iter_rows())
            for row in rows:
                if row[0].value == 'version':
                    version = row[1].value
            self.version = version
        except:
            logger.info(f'could not find a sheet with name "metadata" in workbook. defaulting to v2')
        return version

    def check_all_groups_always_present(self, definitions_list):
        """
        check all variables have the same set of groupings and that it is the same set as the yaml file dictates
        todo: this might not work for countries not listed in the yaml, write a test or more experimenting?
        :param definitions_list: The definitions dictionary generated by ref_date_handling
        :param groups: Contains list of groups used by the model
        :return:
        """

        groups = None

        for variable in definitions_list:
            for value in variable.values():
                if isinstance(value, dict):
                    if groups is not None:
                        assert list(value.keys()) == groups, \
                            f"Expected values for groups: {groups}, but got values for groups: {list(value.keys())}"
                    else:
                        groups = list(value.keys())

    def assert_workbook_valid(self, workbook: openpyxl.Workbook, **kwargs):
        """
        Assert that a workbook is fully valid. Parses each sheet individually.
        :param workbook: The openpyxl workbook object.
        :return:
        """
        has_primary_sheet: bool = False

        for sheetname in workbook.sheetnames:
            rows = list(workbook[sheetname].iter_rows())
            if len(rows) == 0 and sheetname not in ['metadata', 'changes']:
                raise TableValidationError(f'Table is missing header row for sheet {sheetname}')

            # the primary sheet has no constricted name, but convention is to call it 'params'
            # it is simply the only sheet that has 'variable' in cell A1.
            if rows[0][0].value == 'variable':
                has_primary_sheet = True
                self.assert_primary_sheet_valid(rows, sheetname)

            # group-level sheets have names that are countrified var names.
            # todo NOTE this is subject to change to work around the 31-char sheet name limit! Will require revision.
            elif rows[0][0].value == 'group':
                self.assert_group_sheet_valid(rows, sheetname, **kwargs)

        if has_primary_sheet is False:
            raise TableValidationError('Table has no primary data sheets')

    def assert_primary_sheet_valid(self, rows, sheetname):
        """
        Assert that a primary sheet is fully valid, both in rows and columns.
        :param rows: List of rows in the sheet to be parsed
        :param sheetname: Name of the sheet to be parsed
        :return:
        """
        header = [cell.value for cell in rows[0]]
        rows = rows[1:]

        self.assert_no_invalid_primary_headers(header, sheetname)
        indices = self.fetch_primary_header_indices(header, sheetname)

        for i, row in enumerate(rows):
            if row[0].value is not None:
                self.assert_primary_row_valid(row, i + 2, indices, sheetname)

    def assert_no_invalid_primary_headers(self, header, sheetname):
        if 'group' in header:
            raise TableValidationError(f'{sheetname} is a primary sheet. It cannot have a \'group\' column.')

    def fetch_primary_header_indices(self, header, sheetname):
        """
        Generates a dict Str->Int of header names to the index of that column in the primary sheet.
        :param header:
        :param sheetname:
        :return:
        """
        indices = {}

        self.fetch_header_index(indices, header, 'type', sheetname)
        self.fetch_header_index(indices, header, 'param', sheetname)
        self.fetch_header_index(indices, header, 'ref value', sheetname)
        self.fetch_header_index(indices, header, 'ref date', sheetname)
        self.fetch_header_index(indices, header, 'mean growth', sheetname)
        self.fetch_header_index(indices, header, 'initial_value_proportional_variation', sheetname)
        self.fetch_header_index(indices, header, 'variability growth', sheetname)
        self.fetch_header_index(indices, header, 'unit', sheetname)
        self.fetch_header_index(indices, header, 'user name', sheetname)
        self.fetch_header_index(indices, header, 'id', sheetname)
        self.fetch_header_index(indices, header, 'order', sheetname)
        self.fetch_header_index(indices, header, 'ui variable', sheetname)

        self.fetch_optional_header_index(indices, header, 'description', warn=True, sheetname=sheetname)

        self.fetch_optional_header_index(indices, header, 'label')
        self.fetch_optional_header_index(indices, header, 'source')
        self.fetch_optional_header_index(indices, header, 'comment')
        self.fetch_optional_header_index(indices, header, 'control')
        self.fetch_optional_header_index(indices, header, 'scenario notes')
        self.fetch_optional_header_index(indices, header, 'override')

        return indices

    def fetch_header_index(self, indices: Dict[str, int], header: List[str], column_header: str, sheetname: str):
        try:
            indices[column_header] = header.index(column_header)
        except ValueError:
            raise TableValidationError(f'Table is missing {column_header} column for sheet {sheetname}')

    def fetch_optional_header_index(self, indices: Dict[str, int], header: List[str], column_header: str,
                                    warn: bool = False, sheetname: str = None):
        try:
            indices[column_header] = header.index(column_header)
        except ValueError:
            if warn:
                logger.warning(f'Table is missing {column_header} column for sheet {sheetname}')
            indices[column_header] = None

    def assert_primary_row_valid(self, row, row_num: int, indices: Dict[str, int], sheetname: int):
        """
        Given a specific row in a primary sheet, check the vals and ensure they are appropriate.
        :param row: Contains the row as a list of vals.
        :param row_num: Row number in the sheet, for error logging.
        :param indices: A dictionary of header names to their index.
        :param sheetname: Name of the primary sheet to be parsed.
        :return:
        """
        if not isinstance(row[0].value, str):
            raise TableValidationError(f'variable on row {row_num} of sheet {sheetname} not a string')

        variable = row[0].value

        var_type = row[indices['type']].value
        if not var_type in ['exp', 'interp']:
            raise TableValidationError(f'type for {variable} on sheet {sheetname} was {var_type}. Must be one of '
                                       f'[\'exp\', \'interp\']')

        param = row[indices['param']].value
        if var_type == 'interp':
            if not param in ['linear']:
                raise TableValidationError(f'param for {variable} on sheet {sheetname} was {param}. Must be one of '
                                           f'[\'linear\']')
        else:
            if param is not None:
                logger.warning(f'param not empty for non-interp variable {variable} on sheet {sheetname}')

        ref_value = row[indices['ref value']].value
        if var_type == 'interp':
            try:
                ref_value_json = json.loads(ref_value)
            except json.JSONDecodeError:
                raise TableValidationError(f'ref value for interp variable {variable} on sheet {sheetname} was '
                                           f'{ref_value}. Must be valid json')
            dates = list(ref_value_json.keys())
            if not len(dates) == 2:
                raise TableValidationError(f'ref value json for interp variable {variable} on sheet {sheetname} must '
                                           f'have two keys')
            try:
                datetime.datetime.strptime(dates[0], '%Y-%m-%d')
                datetime.datetime.strptime(dates[1], '%Y-%m-%d')
            except ValueError:
                raise TableValidationError(f'ref value json for interp variable {variable} on sheet {sheetname} must '
                                           f'have date keys, in the format YYYY-MM-DD')

            if not isinstance(ref_value_json[dates[0]], Number) or \
                not isinstance(ref_value_json[dates[1]], Number):
                raise TableValidationError(f'ref value json for interp variable {variable} on sheet {sheetname} must '
                                           f'have numeric values')

    def assert_group_sheet_valid(self, rows, sheetname, **kwargs):
        """
        Make sure a given group sheet is also valid.
        Each necessary column must be present, and each row must have correctly formatted corresponding values

        This is fairly hacky since structure of group pages is subject to change with aliasing etc.
        :param rows: List of rows in the sheet to be parsed
        :param sheetname: Name of the group sheet
        :return:
        """
        header = [cell.value for cell in rows[0]]
        rows = rows[1:]

        minimal_viable_header = ['group', 'scenario', 'ref value', 'mean growth',
                           'initial_value_proportional_variation', 'variability growth', 'id']

        for column in minimal_viable_header:
            if column not in header:
                raise TableValidationError(f'Missing header column {column} in group sheet {sheetname}')

        index_column_map = {header.index(h): h for h in header}

        for i, row in enumerate(rows):
            if row[0].value is not None:
                self.assert_group_row_valid(row, sheetname, index_column_map, **kwargs)

    def assert_group_row_valid(self, row, sheetname, index_column_map, **kwargs):
        pass

    def load_definitions(self, sheet_name=None, filename: str = None, **kwargs):
        """
        Loads definitions from the excel workbook
        If sheet_name is given only that sheet will be parsed; if left as None, all sheets will be used.

        :param sheet_name: The name of the sheet to be used; if left blank, all sheets used instead.
        :param filename: The workbook to be parsed
        :param id_flag: Whether missing ids in the excel book should be assigned
        :return: A list of dictionaries containing all the variable value data
        """
        from openpyxl import load_workbook
        wb = load_workbook(filename, data_only=True)

        self.assert_workbook_valid(wb, **kwargs)

        # maps variables to their scenario and group-specific values.
        inline_groupings = {}
        # maps variables to their values, but has dictionaries for each value with different group values
        definitions = defaultdict(lambda: defaultdict(dict))

        _sheet_names = [sheet_name] if sheet_name else wb.sheetnames
        version = self.get_version(wb)

        table_visitor_partial = partial(self.table_visitor, wb=wb, sheet_names=_sheet_names,
                                        definitions=definitions, group_variables=inline_groupings, **kwargs)

        # the first visitor pass is for groups, to build the inline_groupings object,
        # the second visitor pass builds the definitions object.
        if kwargs.get('with_group'):
            table_visitor_partial(visitor_function=self.group_builder)
        table_visitor_partial(visitor_function=self.build_definitions)

        definitions_list = []
        for var_set in definitions.values():
            for scenario_var in var_set.values():
                definitions_list.append(scenario_var)

        self.check_all_groups_always_present(definitions_list)
        return definitions_list

    def correct_ids(self, filename):
        # Handles id logic and generates an id_map dictionary
        from table_data_reader import id_handler
        id_map, highest_id = id_handler.build_id_dict(filename)
        if id_handler.check_for_duplicate_ids(id_map):
            raise Exception("Duplicate ID variable found")
        id_handler.fill_missing_ids(filename, id_map, highest_id)


class XLWingsTableHandler(TableHandler):
    @staticmethod
    def get_sheet_range_bounds(filename, sheet_name):
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows())
        return len(rows)

    def load_definitions(self, sheet_name, filename=None, id_flag=False):
        import xlwings as xw
        definitions = []
        wb = xw.Book(fullname=filename)
        _sheet_names = [sheet_name] if sheet_name else wb.sheets
        for _sheet_name in _sheet_names:
            sheet = wb.sheets[_sheet_name]
            range = sheet.range('A1').expand()
            rows = range.rows
            header = [cell.value for cell in rows[0]]

            # check if this sheet contains parameters or if it documentation
            if header[0] != 'variable':
                continue

            total_rows = self.get_sheet_range_bounds(filename, _sheet_name)
            range = sheet.range((1, 1), (total_rows, len(header)))
            rows = range.rows
            for row in rows[1:]:
                values = {}
                for key, cell in zip(header, row):
                    values[key] = cell.value
                definitions.append(values)
        return definitions


class TableParameterLoader(object):
    definition_version: int
    """Utility to populate ParameterRepository from spreadsheets.

        The structure of the spreadsheets is:

        | variable | ... |
        |----------|-----|
        |   ...    | ... |

        If the first row in a spreadsheet does not contain they keyword 'variable' the sheet is ignored.

       """

    def __init__(self, filename, table_handler='openpyxl', version=2, **kwargs):
        self.filename = filename
        self.definition_version = 2  # default - will be overwritten by handler

        logger.info(f'Using {table_handler} excel handler')
        table_handler_instance = None
        if table_handler == 'csv':
            table_handler_instance = CSVHandler(version)
        if table_handler == 'pandas':
            table_handler_instance = PandasCSVHandler(version)
        if table_handler == 'openpyxl':
            table_handler_instance = OpenpyxlTableHandler()
        if table_handler == 'xlsx2csv':
            table_handler_instance = Xlsx2CsvHandler()
        if table_handler == 'xlwings':
            table_handler_instance = XLWingsTableHandler()
        self.table_handler: TableHandler = table_handler_instance

    def load_parameter_definitions(self, sheet_name: str = None, **kwargs):
        """
        Load variable text from rows in excel file.
        If no spreadsheet arg is given, all spreadsheets are loaded.
        The first cell in the first row in a spreadsheet must contain the keyword 'variable' or the sheet is ignored.

        Any cells used as titles (with no associated value) are also added to the returned dictionary. However, the
        values associated with each header will be None. For example, given the spreadsheet:

        | variable | A | B |
        |----------|---|---|
        | Title    |   |   |
        | Entry    | 1 | 2 |

        The following list of definitions would be returned:

        [ { variable: 'Title', A: None, B: None }
        , { variable: 'Entry', A: 1   , B: 2    }
        ]

        :param sheet_name:
        :return: list of dicts with {header col name : cell value} pairs
        """
        definitions = self.table_handler.load_definitions(sheet_name, filename=self.filename, **kwargs)
        self.definition_version = self.table_handler.version
        return definitions

    def load_into_repo(self, repository: ParameterRepository = None, sheet_name: str = None, **kwargs):
        """
        Create a Repo from an excel file.
        :param repository: the repository to load into
        :param sheet_name:
        :return:
        """
        repository.add_all(self.load_parameters(sheet_name, **kwargs))

    def load_parameters(self, sheet_name, **kwargs):
        # todo: work out what is going on here and comment it

        # load the data in from the spreadsheet (perfectly good, useable data)
        parameter_definitions = self.load_parameter_definitions(sheet_name=sheet_name, **kwargs)
        params = []

        # get a list of parameter names from some weird hardcoded thing
        # todo: ask why
        param_name_map = param_name_maps[int(self.definition_version)]

        for _def in parameter_definitions:
            # substitute names from the headers with the kwargs names in the Parameter and Distributions classes
            # e.g. 'variable' -> 'name', 'module' -> 'module_name', etc
            parameter_kwargs_def = {}
            for k, v in _def.items():
                if k in param_name_map:
                    if param_name_map[k]:
                        parameter_kwargs_def[param_name_map[k]] = v
                    else:
                        parameter_kwargs_def[k] = v
                elif kwargs.get('with_group') and k in kwargs['group_vars']:
                    parameter_kwargs_def[k] = {}
                    for l, w in _def[k].items():
                        if l in param_name_map:
                            if param_name_map[l]:
                                parameter_kwargs_def[k][param_name_map[l]] = w
                            else:
                                parameter_kwargs_def[k][l] = w
            name_ = parameter_kwargs_def['name']
            del parameter_kwargs_def['name']
            p = Parameter(name_, version=self.definition_version, **parameter_kwargs_def)
            params.append(p)
        return params
