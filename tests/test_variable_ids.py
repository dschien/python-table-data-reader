import unittest
import openpyxl
import pytest

import table_data_reader
from deepdiff import DeepDiff
import os
import re

from table_data_reader.table_handlers import OpenpyxlTableHandler

from contextlib import contextmanager


def create_temp_copy(original):
    wb = openpyxl.load_workbook(original, data_only=True)
    wb_copy = openpyxl.Workbook()
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        sheet_a = wb_copy.create_sheet(sheet_name)
        sheet_b = wb[sheet_name]
        for row in range(len(list(sheet_b.iter_rows()))):
            for col in range(len(list(sheet_b.iter_cols()))):
                c = sheet_a.cell(row=row + 1, column=col + 1)
                val = sheet_b.cell(row=row + 1, column=col + 1).value
                c.value = val

    std = wb_copy['Sheet']
    wb_copy.remove(std)
    name = original.split('.')[0] + "_copy." + original.split('.')[1]
    wb_copy.save(name)
    return name


def get_diff(a, b):
    wb_a = openpyxl.load_workbook(a, data_only=True)
    wb_b = openpyxl.load_workbook(b, data_only=True)
    reg1 = re.compile(r".*style.*")
    reg2 = re.compile(r"iterable_item_added")
    diff = DeepDiff(wb_a, wb_b, exclude_regex_paths=[reg1, reg2])
    return diff


def delete_temp_copy(copy):
    os.remove(copy)


def get_static_path(filename):
    """
    Direct copy of the function in eam-core-provenance/tests/directory_test_controller.py
    Get the current script directory- which should point to /tests- and join it with the desired filename, then return
    """
    directory = os.path.dirname(os.path.realpath(__file__))
    return os.path.join(directory, filename)


@contextmanager
def use_copy_of(file):
    copy = create_temp_copy(file)

    try:
        yield
    finally:
        delete_temp_copy(copy)


class TestVariableIDs(unittest.TestCase):

    def test_existing_ids(self):
        with use_copy_of(get_static_path('data/existing_ids.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params",
                                     filename=get_static_path('data/existing_ids_copy.xlsx'),
                                     id_flag=True)

            diff = get_diff(get_static_path('data/existing_ids.xlsx'),
                            get_static_path('data/existing_ids_copy.xlsx'))

        assert 'values_changed' not in diff.keys() and 'type_changes' not in diff.keys()

    def test_existing_ids_group(self):
        with use_copy_of(get_static_path('data/existing_ids_group.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params",
                                     filename=get_static_path('data/existing_ids_group_copy.xlsx'),
                                     id_flag=True,
                                     with_group=True,
                                     groupings=['UK', 'DE'],
                                     group_vars=['power_laptop', 'energy_intensity_network'])

            diff = get_diff(get_static_path('data/existing_ids_group.xlsx'),
                            get_static_path('data/existing_ids_group_copy.xlsx'))

        assert 'values_changed' not in diff.keys() and 'type_changes' not in diff.keys()

    def test_some_existing_ids(self):
        with use_copy_of(get_static_path('data/some_existing_ids.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params",
                                     filename=get_static_path('data/some_existing_ids_copy.xlsx'),
                                     id_flag=True)

            diff = get_diff(get_static_path('data/some_existing_ids.xlsx'),
                            get_static_path('data/some_existing_ids_copy.xlsx'))

        assert diff['type_changes']['root[0][2][18]._value']['new_value'] == 3

    def test_some_existing_ids_group(self):
        with use_copy_of(get_static_path('data/some_existing_ids_group.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params",
                                     filename=get_static_path('data/some_existing_ids_group_copy.xlsx'),
                                     id_flag=True,
                                     with_group=True,
                                     groupings=['UK', 'DE'],
                                     group_vars=['power_laptop', 'energy_intensity_network'])

            diff = get_diff(get_static_path('data/some_existing_ids_group.xlsx'),
                            get_static_path('data/some_existing_ids_group_copy.xlsx'))

        assert diff['type_changes']['root[0][2][19]._value']['new_value'] == 5
        assert diff['type_changes']['root[0][4][19]._value']['new_value'] == 6
        assert diff['type_changes']['root[1][2][6]._value']['new_value'] == 7

    def test_no_existing_ids(self):
        with use_copy_of(get_static_path('data/no_existing_ids.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params",
                                     filename=get_static_path('data/no_existing_ids_copy.xlsx'),
                                     id_flag=True)

            diff = get_diff(get_static_path('data/no_existing_ids.xlsx'),
                            get_static_path('data/no_existing_ids_copy.xlsx'))

        assert diff['type_changes']['root[0][1][18]._value']['new_value'] == 0
        assert diff['type_changes']['root[0][2][18]._value']['new_value'] == 1

    def test_no_existing_ids_group(self):
        with use_copy_of(get_static_path('data/no_existing_ids_group.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params",
                                     filename=get_static_path('data/no_existing_ids_group_copy.xlsx'),
                                     id_flag=True,
                                     with_group=True,
                                     groupings=['UK', 'DE'],
                                     group_vars=['power_laptop', 'energy_intensity_network'])

            diff = get_diff(get_static_path('data/no_existing_ids_group.xlsx'),
                            get_static_path('data/no_existing_ids_group_copy.xlsx'))

        assert diff['type_changes']['root[0][1][19]._value']['new_value'] == 0
        assert diff['type_changes']['root[0][2][19]._value']['new_value'] == 1
        assert diff['type_changes']['root[0][3][19]._value']['new_value'] == 2
        assert diff['type_changes']['root[0][4][19]._value']['new_value'] == 3
        assert diff['type_changes']['root[1][1][6]._value']['new_value'] == 4
        assert diff['type_changes']['root[1][2][6]._value']['new_value'] == 5

    def test_duplicate_ids(self):
        with self.assertRaises(Exception) as context:
            with use_copy_of(get_static_path('data/duplicate_ids.xlsx')):
                handler = OpenpyxlTableHandler()
                handler.load_definitions("params", filename=get_static_path('data/duplicate_ids_copy.xlsx'), id_flag=True)

        self.assertTrue("Duplicate ID variable " in str(context.exception))

    def test_duplicate_ids_group(self):
        with self.assertRaises(Exception) as context:
            with use_copy_of(get_static_path('data/duplicate_ids_group.xlsx')):
                handler = OpenpyxlTableHandler()
                handler.load_definitions("params",
                                         filename=get_static_path('data/duplicate_ids_group_copy.xlsx'),
                                         id_flag=True,
                                         with_group=True,
                                         groupings=['UK', 'DE'],
                                         group_vars=['power_laptop', 'energy_intensity_network'])

        self.assertTrue("Duplicate ID variable " in str(context.exception))

    def test_no_id_flag(self):
        with use_copy_of(get_static_path('data/no_existing_ids.xlsx')):
            handler = OpenpyxlTableHandler()
            handler.load_definitions("params", filename=get_static_path('data/no_existing_ids_copy.xlsx'))

            diff = get_diff(get_static_path('data/no_existing_ids.xlsx'),
                            get_static_path('data/no_existing_ids_copy.xlsx'))

        assert 'values_changed' not in diff.keys() and 'type_changes' not in diff.keys()

    def test_no_id_column(self):
        with self.assertRaises(Exception) as context:
            with use_copy_of(get_static_path('data/no_id_column.xlsx')):
                handler = OpenpyxlTableHandler()
                handler.load_definitions("params",
                                         filename=get_static_path('data/no_id_column_copy.xlsx'),
                                         id_flag=True)

        self.assertTrue(" has no id column" in str(context.exception),
                        f"' has no id column' not found in {str(context.exception)}")

    def test_no_id_column_group(self):
        with self.assertRaises(Exception) as context:
            with use_copy_of(get_static_path('data/no_id_column_group.xlsx')):
                handler = OpenpyxlTableHandler()
                handler.load_definitions("params",
                                         filename=get_static_path('data/no_id_column_group_copy.xlsx'),
                                         id_flag=True,
                                         with_group=True,
                                         groupings=['UK', 'DE'],
                                         group_vars=['power_laptop', 'energy_intensity_network'])

        self.assertTrue(" has no id column" in str(context.exception),
                        f"' has no id column' not found in {str(context.exception)}")

    def test_multiple_id_columns(self):
        with self.assertRaises(Exception) as context:
            with use_copy_of(get_static_path('data/multiple_id_columns.xlsx')):
                handler = OpenpyxlTableHandler()
                handler.load_definitions("params",
                                         filename=get_static_path('data/multiple_id_columns_copy.xlsx'),
                                         id_flag=True)

        self.assertTrue(" has multiple id columns" in str(context.exception),
                        f"' has multiple id columns' not found in {str(context.exception)}")

    def test_multiple_id_columns_group(self):
        with self.assertRaises(Exception) as context:
            with use_copy_of(get_static_path('data/multiple_id_columns_group.xlsx')):
                handler = OpenpyxlTableHandler()
                handler.load_definitions("params",
                                         filename=get_static_path('data/multiple_id_columns_group_copy.xlsx'),
                                         id_flag=True,
                                         with_group=True,
                                         groupings=['UK', 'DE'],
                                         group_vars=['power_laptop', 'energy_intensity_network'])

        self.assertTrue(" has multiple id columns" in str(context.exception),
                        f"' has multiple id columns' not found in {str(context.exception)}")
